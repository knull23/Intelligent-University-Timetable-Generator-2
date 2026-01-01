from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import datetime
import re
from .models import Class, MeetingTime
def abbreviate_course_name(course_name):
    """Abbreviate course name by taking first letter of each word."""
    if not course_name:
        return "UNK"  # Unknown if no name
    words = re.findall(r'\b\w+\b', course_name)
    if not words:
        return course_name.upper()[:3]  # Fallback to first 3 chars
    return ''.join(word[0].upper() for word in words)

def export_timetable_pdf(timetable):
    """Export timetable to PDF format matching TimetableGrid view (Monday-Friday only)."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=36, leftMargin=36,
                           topMargin=36, bottomMargin=18)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        fontSize=14,
        alignment=1,
        spaceAfter=8
    )
    section_title_style = ParagraphStyle(
        'SectionTitleStyle',
        parent=styles['Heading2'],
        fontSize=12,
        alignment=1,
        spaceAfter=6
    )

    story = []
    title = Paragraph(f"{timetable.name}", title_style)
    story.append(title)
    story.append(Spacer(1, 6))

    # Get classes
    classes = timetable.classes.all().select_related('course', 'instructor', 'room', 'meeting_time', 'section')

    # Days: Monday-Friday only
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

    # Fixed time slots matching TimetableGrid
    standard_time_slots = [
        '09:00:00-10:00:00',
        '10:00:00-11:00:00',
        '11:00:00-12:00:00',
        '12:00:00-13:00:00',
        '13:00:00-13:45:00',  # Lunch break
        '13:45:00-14:45:00',
        '14:45:00-15:45:00',
        '15:45:00-16:45:00'
    ]

    # Group classes by section
    sections = {}
    section_rooms = {}
    for cls in classes:
        section_id = cls.section.section_id
        if section_id not in sections:
            sections[section_id] = []
            section_rooms[section_id] = set()
        sections[section_id].append(cls)
        section_rooms[section_id].add(cls.room.room_number)

    # Build course legend for instructor table
    legend_map = {}
    for cls in classes:
        key = abbreviate_course_name(cls.course.course_name)
        if key not in legend_map:
            legend_map[key] = {'course_short': key, 'course': cls.course.course_name, 'instructors': set()}
        legend_map[key]['instructors'].add(cls.instructor.name)

    course_legend = []
    for item in legend_map.values():
        course_legend.append({
            'course_short': item['course_short'],
            'course': item['course'],
            'instructors': list(item['instructors'])
        })

    # For each section, create a timetable
    for section_id, section_classes in sections.items():
        # Section title with room numbers
        room_numbers = ', '.join(sorted(section_rooms[section_id]))
        section_title = Paragraph(f"Timetable for {section_id} (Room: {room_numbers})", section_title_style)
        story.append(section_title)
        story.append(Spacer(1, 6))

        # Build schedule data for this section
        schedule = {day: {} for day in days}
        for cls in section_classes:
            day = cls.meeting_time.day
            if day not in days:
                continue
            # Calculate actual end time based on course duration
            duration_hours = getattr(cls.course, 'duration', 1)
            start_time = cls.meeting_time.start_time

            # For multi-hour classes, split into individual 1-hour slots
            for i in range(duration_hours):
                slot_start = datetime.time(hour=(start_time.hour + i) % 24, minute=start_time.minute)
                slot_end = datetime.time(hour=(slot_start.hour + 1) % 24, minute=slot_start.minute)
                time_slot = f"{slot_start}-{slot_end}"
                is_start = (i == 0)
                schedule.setdefault(day, {})
                schedule[day].setdefault(time_slot, [])
                schedule[day][time_slot].append({
                    'course': cls.course.course_name,
                    'course_id': cls.course.course_id,
                    'instructor': cls.instructor.name,
                    'room': cls.room.room_number,
                    'section': cls.section.section_id,
                    'course_type': cls.course.course_type,
                    'duration': duration_hours,
                    'is_start': is_start,
                    'colspan': duration_hours if is_start else 1
                })

        # Build schedule table for this section
        table_data = []
        header_row = ['Day'] + [f"{slot.split('-')[0][:5]}-{slot.split('-')[1][:5]}" + (" (LUNCH)" if slot == '13:00:00-13:45:00' else "") for slot in standard_time_slots]
        table_data.append(header_row)

        for day in days:
            row = [day]
            for slot_index, time_slot in enumerate(standard_time_slots):
                if time_slot == '13:00:00-13:45:00':
                    # Lunch break
                    row.append("LUNCH BREAK")
                    continue

                day_classes = schedule[day].get(time_slot, [])

                # Check if this slot should be skipped due to colspan from previous multi-hour class
                skip_slot = False
                for prev_index in range(slot_index):
                    prev_slot = standard_time_slots[prev_index]
                    prev_classes = schedule[day].get(prev_slot, [])
                    for prev_cls in prev_classes:
                        if prev_cls['is_start'] and prev_cls['colspan'] > 1:
                            start_index = standard_time_slots.index(prev_slot)
                            end_index = start_index + prev_cls['colspan'] - 1
                            if slot_index <= end_index:
                                skip_slot = True
                                break
                    if skip_slot:
                        break

                if skip_slot:
                    row.append("")
                    continue

                # Check if there's a multi-hour class starting at this slot
                multi_hour_class = next((cls for cls in day_classes if cls['is_start'] and cls['colspan'] > 1), None)

                cell_content_lines = []
                for cls in day_classes:
                    course_short = abbreviate_course_name(cls['course'])  # Use abbreviated course name
                    instructor = cls['instructor']
                    cell_content_lines.append(f"{course_short}")
                    cell_content_lines.append(f"{instructor}")
                    if multi_hour_class and cls['is_start']:
                        start_slot = time_slot
                        end_slot = standard_time_slots[slot_index + cls['colspan'] - 1]
                        cell_content_lines.append(f"({start_slot.split('-')[0][:5]} - {end_slot.split('-')[1][:5]})")
                    cell_content_lines.append("")

                cell_content = "\n".join(cell_content_lines).strip()
                row.append(cell_content)

            table_data.append(row)

        if len(table_data) == 1:
            # No slots for this section
            story.append(Paragraph(f"No scheduled classes for {section_id}.", styles['Normal']))
        else:
            # Create schedule table
            col_widths = [60] + [80] * len(standard_time_slots)  # Adjust column widths
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f4f4f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))

            story.append(table)

        story.append(Spacer(1, 12))
        story.append(PageBreak())  # Page break after each section

    # Instructor Information Table at the end
    if course_legend:
        instructor_title = Paragraph("Instructor Information", title_style)
        story.append(instructor_title)
        story.append(Spacer(1, 6))

        instructor_data = [['Instructor', 'Courses & Short Forms']]
        for item in course_legend:
            instructors_str = ', '.join(item['instructors'])
            course_str = f"{item['course_short']} - {item['course']}"
            instructor_data.append([instructors_str, course_str])

        instructor_table = Table(instructor_data, colWidths=[120, 300])
        instructor_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f4f4f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        story.append(instructor_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def export_timetable_excel(timetable):
    """Export timetable to Excel (Monday-Friday)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    classes = timetable.classes.all().select_related('course', 'instructor', 'room', 'meeting_time', 'section')

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

    # Fixed time slots matching TimetableGrid
    standard_time_slots = [
        '09:00:00-10:00:00',
        '10:00:00-11:00:00',
        '11:00:00-12:00:00',
        '12:00:00-13:00:00',
        '13:00:00-13:45:00',  # Lunch break
        '13:45:00-14:45:00',
        '14:45:00-15:45:00',
        '15:45:00-16:45:00'
    ]

    # Group classes by section
    sections = {}
    section_rooms = {}
    for cls in classes:
        section_id = cls.section.section_id
        if section_id not in sections:
            sections[section_id] = []
            section_rooms[section_id] = set()
        sections[section_id].append(cls)
        section_rooms[section_id].add(cls.room.room_number)

    # Build course legend for instructor table
    legend_map = {}
    for cls in classes:
        key = abbreviate_course_name(cls.course.course_name)
        if key not in legend_map:
            legend_map[key] = {'course_short': key, 'course': cls.course.course_name, 'instructors': set()}
        legend_map[key]['instructors'].add(cls.instructor.name)

    course_legend = []
    for item in legend_map.values():
        course_legend.append({
            'course_short': item['course_short'],
            'course': item['course'],
            'instructors': list(item['instructors'])
        })

    # For each section, create a timetable
    current_row = 1
    for section_id, section_classes in sections.items():
        # Section title with room numbers
        room_numbers = ', '.join(sorted(section_rooms[section_id]))
        ws.cell(row=current_row, column=1, value=f"Timetable for {section_id} (Room: {room_numbers})")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
        ws.cell(row=current_row, column=1).alignment = center_alignment
        current_row += 2

        # Build schedule data for this section
        schedule = {day: {} for day in days}
        for cls in section_classes:
            day = cls.meeting_time.day
            if day not in days:
                continue
            # Calculate actual end time based on course duration
            duration_hours = getattr(cls.course, 'duration', 1)
            start_time = cls.meeting_time.start_time

            # For multi-hour classes, split into individual 1-hour slots
            for i in range(duration_hours):
                slot_start = datetime.time(hour=(start_time.hour + i) % 24, minute=start_time.minute)
                slot_end = datetime.time(hour=(slot_start.hour + 1) % 24, minute=slot_start.minute)
                time_slot = f"{slot_start}-{slot_end}"
                is_start = (i == 0)
                schedule.setdefault(day, {})
                schedule[day].setdefault(time_slot, [])
                schedule[day][time_slot].append({
                    'course': cls.course.course_name,
                    'course_id': cls.course.course_id,
                    'instructor': cls.instructor.name,
                    'room': cls.room.room_number,
                    'section': cls.section.section_id,
                    'course_type': cls.course.course_type,
                    'duration': duration_hours,
                    'is_start': is_start,
                    'colspan': duration_hours if is_start else 1
                })

        # Build schedule table for this section
        headers = ['Time'] + days
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        current_row += 1

        for slot_index, time_slot in enumerate(standard_time_slots):
            row = [f"{time_slot.split('-')[0][:5]}-{time_slot.split('-')[1][:5]}" + (" (LUNCH)" if time_slot == '13:00:00-13:45:00' else "")]
            for day in days:
                if time_slot == '13:00:00-13:45:00':
                    # Lunch break
                    row.append("LUNCH BREAK")
                    continue

                day_classes = schedule[day].get(time_slot, [])

                # Check if this slot should be skipped due to colspan from previous multi-hour class
                skip_slot = False
                for prev_index in range(slot_index):
                    prev_slot = standard_time_slots[prev_index]
                    prev_classes = schedule[day].get(prev_slot, [])
                    for prev_cls in prev_classes:
                        if prev_cls['is_start'] and prev_cls['colspan'] > 1:
                            start_index = standard_time_slots.index(prev_slot)
                            end_index = start_index + prev_cls['colspan'] - 1
                            if slot_index <= end_index:
                                skip_slot = True
                                break
                    if skip_slot:
                        break

                if skip_slot:
                    row.append("")
                    continue

                # Check if there's a multi-hour class starting at this slot
                multi_hour_class = next((cls for cls in day_classes if cls['is_start'] and cls['colspan'] > 1), None)

                cell_content_lines = []
                for cls in day_classes:
                    course_short = abbreviate_course_name(cls['course'])  # Use abbreviated course name
                    instructor = cls['instructor']
                    cell_content_lines.append(f"{course_short}")
                    cell_content_lines.append(f"{instructor}")
                    if multi_hour_class and cls['is_start']:
                        start_slot = time_slot
                        end_slot = standard_time_slots[slot_index + cls['colspan'] - 1]
                        cell_content_lines.append(f"({start_slot.split('-')[0][:5]} - {end_slot.split('-')[1][:5]})")
                    cell_content_lines.append("")

                cell_content = "\n".join(cell_content_lines).strip()
                row.append(cell_content)

            for col, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.alignment = center_alignment
            current_row += 1

        current_row += 2  # Space between sections

    # Instructor Information Table at the end
    if course_legend:
        ws.cell(row=current_row, column=1, value="Instructor Information")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
        ws.cell(row=current_row, column=1).alignment = center_alignment
        current_row += 1

        instructor_headers = ['Instructor', 'Courses & Short Forms']
        for col, header in enumerate(instructor_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        current_row += 1

        for item in course_legend:
            instructors_str = ', '.join(item['instructors'])
            course_str = f"{item['course_short']} - {item['course']}"
            ws.cell(row=current_row, column=1, value=instructors_str)
            ws.cell(row=current_row, column=2, value=course_str)
            current_row += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def check_instructor_conflicts(timetable, instructor_id, new_day, new_start_time, exclude_class_id=None):
    """
    Check if moving an instructor to a new slot would create conflicts.
    Returns a list of conflicting classes with their details.
    """
    conflicts = []

    # Get all classes for this instructor in the timetable (excluding the one being moved)
    instructor_classes = timetable.classes.filter(
        instructor_id=instructor_id
    ).exclude(class_id=exclude_class_id).select_related('meeting_time', 'section', 'course')

    for cls in instructor_classes:
        # Check if the new slot overlaps with any existing class for this instructor
        if cls.meeting_time.day == new_day:
            # For multi-hour classes, check all slots they occupy
            duration = getattr(cls.course, 'duration', 1)
            for hour_offset in range(duration):
                existing_start = datetime.time(
                    hour=(cls.meeting_time.start_time.hour + hour_offset) % 24,
                    minute=cls.meeting_time.start_time.minute
                )

                # Check if this conflicts with the new slot
                if existing_start == new_start_time:
                    conflicts.append({
                        'day': cls.meeting_time.day,
                        'time': f"{cls.meeting_time.start_time.strftime('%H:%M')}-{cls.meeting_time.end_time.strftime('%H:%M')}",
                        'section': cls.section.section_id,
                        'course': cls.course.course_name,
                        'room': cls.room.room_number
                    })
                    break  # No need to check further hours for this class

    return conflicts


def check_slot_conflicts(timetable, new_day, new_start_time, instructor_id, room_id, section_id, exclude_class_id=None):
    """
    Check for conflicts when moving a class to a new slot.
    Checks for instructor, room, and section conflicts.
    Returns a list of conflicting classes with their details.
    """
    conflicts = []

    # Get all classes in the timetable (excluding the one being moved)
    all_classes = timetable.classes.filter().exclude(class_id=exclude_class_id).select_related('meeting_time', 'section', 'course', 'instructor', 'room')

    for cls in all_classes:
        # Check if the new slot overlaps with any existing class
        if cls.meeting_time.day == new_day:
            # For multi-hour classes, check all slots they occupy
            duration = getattr(cls.course, 'duration', 1)
            for hour_offset in range(duration):
                existing_start = datetime.time(
                    hour=(cls.meeting_time.start_time.hour + hour_offset) % 24,
                    minute=cls.meeting_time.start_time.minute
                )

                # Check if this conflicts with the new slot
                if existing_start == new_start_time:
                    conflict_type = None

                    # Check for instructor conflict
                    if cls.instructor_id == instructor_id:
                        conflict_type = 'instructor'
                    # Check for room conflict
                    elif cls.room_id == room_id:
                        conflict_type = 'room'
                    # Check for section conflict
                    elif cls.section_id == section_id:
                        conflict_type = 'section'

                    if conflict_type:
                        conflicts.append({
                            'type': conflict_type,
                            'day': cls.meeting_time.day,
                            'time': f"{cls.meeting_time.start_time.strftime('%H:%M')}-{cls.meeting_time.end_time.strftime('%H:%M')}",
                            'section': cls.section.section_id,
                            'course': cls.course.course_name,
                            'room': cls.room.room_number,
                            'instructor': cls.instructor.name
                        })
                    break  # No need to check further hours for this class

    return conflicts
